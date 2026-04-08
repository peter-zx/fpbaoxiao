# -*- coding: utf-8 -*-
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\baoxiao\2026年启用报销与费用填写.xlsx')
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== Sheet: {name} ===')
    for row in ws.iter_rows(max_row=30, values_only=True):
        if any(cell is not None for cell in row):
            print(row)
