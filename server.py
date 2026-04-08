# -*- coding: utf-8 -*-
"""
报销费用填写工具 - 本地服务器
支持图片嵌入Excel
"""

import os
import sys
import json
import base64
import uuid
from datetime import datetime
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import parse_qs, urlparse
import threading
import webbrowser

# 尝试导入openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告: openpyxl未安装，图片嵌入功能不可用")
    print("请运行: pip install openpyxl Pillow")

# 尝试导入Pillow处理图片
try:
    from PIL import Image as PILImage
    import io
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# 数据存储
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
IMAGE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'images')

# 确保图片目录存在
os.makedirs(IMAGE_DIR, exist_ok=True)

def load_data():
    """加载保存的数据"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {'expense': [], 'reimburse': []}

def save_data(data):
    """保存数据"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def save_image(base64_data, record_id, field_name):
    """保存base64图片到本地文件，返回PIL Image对象"""
    try:
        # 解析base64数据
        if ',' in base64_data:
            header, data = base64_data.split(',', 1)
            # 获取图片格式
            if 'png' in header:
                ext = 'png'
            elif 'jpeg' in header or 'jpg' in header:
                ext = 'jpg'
            elif 'gif' in header:
                ext = 'gif'
            else:
                ext = 'png'
        else:
            data = base64_data
            ext = 'png'
        
        # 解码
        img_data = base64.b64decode(data)
        
        # 直接用PIL打开并返回
        pil_img = PILImage.open(io.BytesIO(img_data))
        return pil_img
    except Exception as e:
        print(f"解析图片失败: {e}")
        return None

def create_excel_with_images(data, output_path):
    """创建包含图片的Excel文件"""
    wb = Workbook()
    
    # 删除默认sheet
    wb.remove(wb.active)
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 创建费用模板sheet
    ws_expense = wb.create_sheet('费用模板（不计入销售成本）')
    expense_headers = ['时间', '产品（购买的物品、服务）', '关联项目', '发生原因', '金额', '详情截图', '是否有票', '开票主体']
    
    # 写入标题行
    ws_expense.merge_cells('A1:H1')
    ws_expense['A1'] = '报销详情'
    ws_expense['A1'].font = Font(bold=True, size=14)
    ws_expense['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头
    for col, header in enumerate(expense_headers, 1):
        cell = ws_expense.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入费用数据
    expense_records = data.get('expense', [])
    for row_idx, record in enumerate(expense_records, 3):
        ws_expense.cell(row=row_idx, column=1, value=record.get('time', '')).border = thin_border
        ws_expense.cell(row=row_idx, column=2, value=record.get('product', '')).border = thin_border
        ws_expense.cell(row=row_idx, column=3, value=record.get('related', '')).border = thin_border
        ws_expense.cell(row=row_idx, column=4, value=record.get('reason', '')).border = thin_border
        ws_expense.cell(row=row_idx, column=5, value=record.get('amount', 0)).border = thin_border
        ws_expense.cell(row=row_idx, column=6, value='').border = thin_border
        ws_expense.cell(row=row_idx, column=7, value=record.get('hasTicket', '')).border = thin_border
        ws_expense.cell(row=row_idx, column=8, value=record.get('ticketEntity', '')).border = thin_border
        
        # 调整行高以容纳图片
        ws_expense.row_dimensions[row_idx].height = 80
        
        # 插入图片
        if record.get('image') and HAS_OPENPYXL and HAS_PIL:
            try:
                pil_img = save_image(record['image'], record['id'], 'expense')
                if pil_img:
                    # 缩小图片
                    pil_img.thumbnail((150, 100))
                    
                    # 保存到内存中的字节流
                    img_buffer = io.BytesIO()
                    # 根据原始格式保存
                    fmt = pil_img.format if pil_img.format else 'PNG'
                    if fmt == 'JPEG':
                        pil_img.save(img_buffer, format='JPEG')
                    else:
                        pil_img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    # 创建XLImage对象
                    xl_img = XLImage(img_buffer)
                    xl_img.anchor = f'F{row_idx}'
                    ws_expense.add_image(xl_img)
            except Exception as e:
                print(f"插入图片失败: {e}")
    
    # 添加合计行（修复循环引用问题）
    if expense_records:
        total_row = len(expense_records) + 3
        ws_expense.cell(row=total_row, column=1, value='合计').font = Font(bold=True)
        # 公式只计算数据行，不包含合计行自己
        ws_expense.cell(row=total_row, column=5, value=f'=SUM(E3:E{total_row-1})').font = Font(bold=True)
    
    # 调整列宽
    col_widths = [12, 30, 20, 20, 12, 20, 10, 20]
    for col, width in enumerate(col_widths, 1):
        ws_expense.column_dimensions[get_column_letter(col)].width = width
    
    # 创建报销模板sheet
    ws_reimburse = wb.create_sheet('报销模板（计入销售订单成本）')
    reimburse_headers = ['时间', '产品（购买的物品、服务）', '关联客户', '发生原因', '金额', '详情截图', '是否有票', '开票主体']
    
    # 写入标题行
    ws_reimburse.merge_cells('A1:H1')
    ws_reimburse['A1'] = '报销详情'
    ws_reimburse['A1'].font = Font(bold=True, size=14)
    ws_reimburse['A1'].alignment = Alignment(horizontal='center')
    
    # 写入表头
    for col, header in enumerate(reimburse_headers, 1):
        cell = ws_reimburse.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入报销数据
    reimburse_records = data.get('reimburse', [])
    for row_idx, record in enumerate(reimburse_records, 3):
        ws_reimburse.cell(row=row_idx, column=1, value=record.get('time', '')).border = thin_border
        ws_reimburse.cell(row=row_idx, column=2, value=record.get('product', '')).border = thin_border
        ws_reimburse.cell(row=row_idx, column=3, value=record.get('related', '')).border = thin_border
        ws_reimburse.cell(row=row_idx, column=4, value=record.get('reason', '')).border = thin_border
        ws_reimburse.cell(row=row_idx, column=5, value=record.get('amount', 0)).border = thin_border
        ws_reimburse.cell(row=row_idx, column=6, value='').border = thin_border
        ws_reimburse.cell(row=row_idx, column=7, value=record.get('hasTicket', '')).border = thin_border
        ws_reimburse.cell(row=row_idx, column=8, value=record.get('ticketEntity', '')).border = thin_border
        
        # 调整行高
        ws_reimburse.row_dimensions[row_idx].height = 80
        
        # 插入图片
        if record.get('image') and HAS_OPENPYXL and HAS_PIL:
            try:
                pil_img = save_image(record['image'], record['id'], 'reimburse')
                if pil_img:
                    pil_img.thumbnail((150, 100))
                    img_buffer = io.BytesIO()
                    fmt = pil_img.format if pil_img.format else 'PNG'
                    if fmt == 'JPEG':
                        pil_img.save(img_buffer, format='JPEG')
                    else:
                        pil_img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    xl_img = XLImage(img_buffer)
                    xl_img.anchor = f'F{row_idx}'
                    ws_reimburse.add_image(xl_img)
            except Exception as e:
                print(f"插入图片失败: {e}")
    
    # 添加合计行（修复循环引用）
    if reimburse_records:
        total_row = len(reimburse_records) + 3
        ws_reimburse.cell(row=total_row, column=1, value='合计').font = Font(bold=True)
        ws_reimburse.cell(row=total_row, column=5, value=f'=SUM(E3:E{total_row-1})').font = Font(bold=True)
    
    # 调整列宽
    for col, width in enumerate(col_widths, 1):
        ws_reimburse.column_dimensions[get_column_letter(col)].width = width
    
    # 保存文件
    wb.save(output_path)
    return output_path


class APIHandler(SimpleHTTPRequestHandler):
    """自定义请求处理器"""
    
    def __init__(self, *args, **kwargs):
        self.directory = os.path.dirname(os.path.abspath(__file__))
        super().__init__(*args, directory=self.directory, **kwargs)
    
    def do_POST(self):
        """处理POST请求"""
        parsed = urlparse(self.path)
        
        if parsed.path == '/api/save':
            # 保存数据
            content_length = int(self.headers['Content-Length'])
            body = self.rfile.read(content_length)
            data = json.loads(body.decode('utf-8'))
            save_data(data)
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())
            
        elif parsed.path == '/api/export':
            # 导出Excel
            content_length = int(self.headers['Content-Length'])
            body = self.rfile.read(content_length)
            data = json.loads(body.decode('utf-8'))
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'报销费用_{timestamp}.xlsx'
            output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
            
            try:
                create_excel_with_images(data, output_path)
                result = {'success': True, 'filename': filename, 'path': output_path}
            except Exception as e:
                import traceback
                traceback.print_exc()
                result = {'success': False, 'error': str(e)}
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(result, ensure_ascii=False).encode())
        
        elif parsed.path == '/api/load':
            # 加载数据
            data = load_data()
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(data, ensure_ascii=False).encode())
        
        else:
            self.send_error(404)
    
    def do_GET(self):
        """处理GET请求"""
        parsed = urlparse(self.path)
        
        if parsed.path == '/api/load':
            # 加载数据
            data = load_data()
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(data, ensure_ascii=False).encode())
        else:
            super().do_GET()
    
    def log_message(self, format, *args):
        """自定义日志格式"""
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {args[0]}")


def main():
    """主函数"""
    port = 8765
    
    # 检查依赖
    if not HAS_OPENPYXL:
        print("\n" + "="*50)
        print("需要安装依赖: pip install openpyxl Pillow")
        print("="*50 + "\n")
        return
    
    print(f"\n{'='*50}")
    print(f"报销费用填写工具 - 本地服务器")
    print(f"{'='*50}")
    print(f"服务地址: http://localhost:{port}")
    print(f"数据文件: {DATA_FILE}")
    print(f"图片目录: {IMAGE_DIR}")
    print(f"{'='*50}\n")
    
    # 启动服务器
    server = HTTPServer(('localhost', port), APIHandler)
    
    # 自动打开浏览器
    def open_browser():
        webbrowser.open(f'http://localhost:{port}')
    
    threading.Timer(1, open_browser).start()
    
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n服务器已停止")
        server.shutdown()


if __name__ == '__main__':
    main()
