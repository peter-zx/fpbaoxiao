# -*- coding: utf-8 -*-
"""
报销费用填写工具 - 桌面版 (server_cloud.py)
===============================================
为打包优化版本：
  1. 端口自动检测（8765被占用自动换下一个）
  2. 单实例检测（防止重复启动）
  3. 自动打开浏览器
  4. 启动时显示二维码（手机扫码）
"""

import os
import sys
import json
import base64
import io
import logging
import signal
import socket
import webbrowser
import threading
import time

# Windows 终端 UTF-8 输出兼容
if sys.platform == 'win32':
    os.system('')  # 激活 ANSI 转义支持
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except AttributeError:
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'replace')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'replace')

from datetime import datetime
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse, quote
from pathlib import Path

# ---- 依赖检查 ----
try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False
    print("❌ 缺少 xlsxwriter，请运行: pip install xlsxwriter")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️ 缺少 openpyxl，Excel导出将使用xlsxwriter")

try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("❌ 缺少 Pillow，请运行: pip install Pillow")

# ---- Office 工具检测 ----
HAS_WIN32COM = False
OFFICE_TYPE = None  # 'excel' 或 'wps'

def detect_office_tool():
    """检测可用的Office工具（Excel优先，其次WPS）"""
    global HAS_WIN32COM, OFFICE_TYPE
    
    # 先尝试 Excel
    try:
        import win32com.client
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Quit()
        HAS_WIN32COM = True
        OFFICE_TYPE = 'excel'
        logging.info("检测到 Microsoft Excel COM")
        return True
    except:
        pass
    
    # 再尝试 WPS ET (WPS表格)
    try:
        import win32com.client
        et = win32com.client.Dispatch('ET.Application')
        et.Quit()
        HAS_WIN32COM = True
        OFFICE_TYPE = 'wps'
        logging.info("检测到 WPS 表格 COM")
        return True
    except:
        pass
    
    # 最后尝试 WPS
    try:
        import win32com.client
        wps = win32com.client.Dispatch('WPS.Application')
        wps.Quit()
        HAS_WIN32COM = True
        OFFICE_TYPE = 'wps'
        logging.info("检测到 WPS COM")
        return True
    except:
        pass
    
    HAS_WIN32COM = False
    OFFICE_TYPE = None
    logging.info("未检测到 Office COM，将使用 xlsxwriter")
    return False


# ==================== 打包兼容处理 ====================
def get_base_dir():
    """获取运行目录（兼容打包后的exe）"""
    if getattr(sys, 'frozen', False):
        # 打包后的exe
        return Path(sys._MEIPASS)
    else:
        # 源代码运行
        return Path(__file__).parent.resolve()


def get_data_dir():
    """获取数据目录（exe运行时在exe旁边创建data文件夹）"""
    if getattr(sys, 'frozen', False):
        # exe同目录下创建data文件夹
        return Path(sys.executable).parent / 'data'
    else:
        # 源代码运行用当前目录
        return Path(__file__).parent.resolve()


# ---- 配置路径 ----
BASE_DIR   = get_base_dir()
DATA_DIR   = get_data_dir()
DATA_FILE  = DATA_DIR / 'data.json'
CONFIG_FILE = DATA_DIR / 'config.json'
LOG_FILE   = DATA_DIR / 'server.log'
OUTPUT_DIR = DATA_DIR / 'exports'
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# 清理旧版遗留的 images 文件夹（新版图片走 base64，不再需要）
import shutil as _shutil
_old_img_dir = Path(__file__).parent.resolve() / 'images'
if _old_img_dir.exists() and _old_img_dir.is_dir():
    try:
        _shutil.rmtree(_old_img_dir)
        print(f"已清理旧版遗留目录: {_old_img_dir}")
    except Exception:
        pass

# ---- 默认配置 ----
DEFAULT_CONFIG = {
    "port":      8765,
    "cors_origins": ["*"],
    "log_level": "INFO",
    "max_content_length": 20 * 1024 * 1024,
}

config = dict(DEFAULT_CONFIG)


def load_config():
    """加载 config.json"""
    global config
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
            config = {**DEFAULT_CONFIG, **loaded}
        except Exception as e:
            config = dict(DEFAULT_CONFIG)
    else:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=2)


def get_local_ip():
    """获取本机局域网IP"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def find_available_port(start_port=8765, max_attempts=100):
    """自动寻找可用端口"""
    for port in range(start_port, start_port + max_attempts):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.bind(('', port))
            s.close()
            return port
        except OSError:
            continue
    return None


def check_port_in_use(port):
    """检查端口是否被占用"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.bind(('', port))
        s.close()
        return False
    except OSError:
        return True


def check_single_instance():
    """单实例检测 - 防止重复启动"""
    import tempfile
    lock_file = DATA_DIR / 'server.lock'
    
    if lock_file.exists():
        # 读取之前的PID
        try:
            with open(lock_file, 'r') as f:
                old_pid = int(f.read().strip())
            
            # 检查进程是否还在 - 尝试终止旧进程
            if sys.platform == 'win32':
                import subprocess
                try:
                    subprocess.run(['taskkill', '/F', '/PID', str(old_pid)], 
                                   capture_output=True, timeout=3)
                    logging.info(f"已终止旧进程 {old_pid}")
                    time.sleep(1)  # 等待端口释放
                except:
                    pass
            
            # 清理锁文件
            lock_file.unlink()
        except:
            pass
    
    # 写入当前PID
    with open(lock_file, 'w') as f:
        f.write(str(os.getpid()))
    return None


# ---- 日志初始化 ----
def init_logging():
    logging.basicConfig(
        level=getattr(logging, config.get('log_level', 'INFO')),
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(LOG_FILE, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )


# ---- 数据存取 ----
def load_data():
    if DATA_FILE.exists():
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.warning(f"加载数据失败: {e}")
    return {'expense': [], 'reimburse': []}


def save_data(data):
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logging.error(f"保存数据失败: {e}")
        return False


def parse_base64_image(base64_data):
    """解析base64图片"""
    if not base64_data:
        return None
    try:
        if ',' in base64_data:
            header, data = base64_data.split(',', 1)
        else:
            data = base64_data
        img_data = base64.b64decode(data)
        pil_img = PILImage.open(io.BytesIO(img_data))
        return pil_img
    except Exception as e:
        logging.error(f"解析图片失败: {e}")
        return None


def prepare_images_xlsx(records, prefix, img_tmp_dir):
    """准备图片，返回 {excel_row: (img_path, pil_img)}"""
    img_map = {}
    for idx, r in enumerate(records):
        if r.get('image'):
            try:
                pil_img = parse_base64_image(r['image'])
                if pil_img:
                    img_path = img_tmp_dir / f'{prefix}_{idx}.png'
                    pil_img.save(str(img_path), format='PNG')
                    # Excel行号：header=1，数据从2开始 → row_idx = idx+2
                    img_map[idx + 2] = (str(img_path), pil_img)
            except Exception as e:
                logging.error(f"保存图片失败: {e}")
    return img_map


def write_sheet_xlsx(wb, ws, records, img_map):
    """用 xlsxwriter 写单个 sheet
    图片用 insert_image() + x_scale/y_scale + object_position=2
    → 缩放后嵌入，随单元格移动（浮动图片，兼容所有 Excel 版本）
    """
    # ---- 表头样式：字号12，加粗，居中，黑色，无背景色 ----
    # 注意：xlsxwriter 的 add_format() 在 workbook 上，不在 worksheet 上
    header_fmt = wb.add_format({
        'bold': True,
        'font_size': 12,
        'font_name': '微软雅黑',
        'font_color': '#000000',
        'bg_color': '#FFFFFF',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
    })
    # ---- 数据行样式 ----
    data_fmt = wb.add_format({
        'font_size': 10,
        'font_name': '微软雅黑',
        'font_color': '#000000',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
    })
    money_fmt = wb.add_format({
        'font_size': 10,
        'font_name': '微软雅黑',
        'font_color': '#000000',
        'align': 'center',
        'valign': 'vcenter',
        'num_format': '¥#,##0.00',
        'border': 1,
    })
    total_label_fmt = wb.add_format({
        'bold': True,
        'font_size': 10,
        'font_name': '微软雅黑',
        'font_color': '#000000',
        'align': 'right',
        'valign': 'vcenter',
        'border': 1,
    })
    total_money_fmt = wb.add_format({
        'bold': True,
        'font_size': 10,
        'font_name': '微软雅黑',
        'font_color': '#000000',
        'align': 'center',
        'valign': 'vcenter',
        'num_format': '¥#,##0.00',
        'border': 1,
    })

    # ---- 列宽：第6列(F)=15（截图列）----
    # xlsxwriter 列宽单位=字符数；1字符≈7.2pt(10pt雅黑)，即 1 char ≈ 9 px (96dpi)
    col_widths = [12, 16, 14, 18, 12, 15, 8, 14]
    ws.set_column(0, 0, col_widths[0])  # A 时间
    ws.set_column(1, 1, col_widths[1])  # B 产品
    ws.set_column(2, 2, col_widths[2])  # C 关联项目
    ws.set_column(3, 3, col_widths[3])  # D 发生原因
    ws.set_column(4, 4, col_widths[4])  # E 金额
    ws.set_column(5, 5, col_widths[5])  # F 截图（较宽）
    ws.set_column(6, 6, col_widths[6])  # G 是否有票
    ws.set_column(7, 7, col_widths[7])  # H 开票主体

    # ---- 表头（无合并行，直接第一行）----
    headers = ['时间', '产品', '关联项目', '发生原因', '金额', '详情截图', '是否有票', '开票主体']
    for col, h in enumerate(headers):
        ws.write(0, col, h, header_fmt)
    ws.set_row(0, 40)  # 表头行高=40

    # ---- 数据行 ----
    total = 0.0
    for idx, r in enumerate(records):
        excel_row = idx + 1   # 0-based row index
        data_row = excel_row + 1  # Excel 1-based 行号（header=0，数据从1开始，合计在最后）

        total += r.get('amount', 0)

        ws.write(excel_row, 0, r.get('time', ''), data_fmt)
        ws.write(excel_row, 1, r.get('product', ''), data_fmt)
        ws.write(excel_row, 2, r.get('related', '') or '-', data_fmt)
        ws.write(excel_row, 3, r.get('reason', ''), data_fmt)
        ws.write_number(excel_row, 4, r.get('amount', 0), money_fmt)
        ws.write(excel_row, 6, r.get('hasTicket', ''), data_fmt)
        ws.write(excel_row, 7, r.get('ticketEntity', '') or '-', data_fmt)

        # 数据行高=200（给图片足够空间）
        ws.set_row(excel_row, 200)

        # 插入图片
        if data_row in img_map:
            img_path, pil_img = img_map[data_row]
            orig_w, orig_h = pil_img.size

            # 等比例缩放：宽高用同一个系数，高度=单元格高度，宽度自动等比
            # xlsxwriter scale=原图像素×系数；row_height=200pt≈267px(96dpi)
            if orig_h > 0:
                scale = 200.0 / orig_h     # 统一系数：高度撑满200pt
                x_scale = scale
                y_scale = scale
            else:
                x_scale = y_scale = 1.0

            ws.insert_image(
                excel_row, 5,
                img_path,
                {
                    'x_scale': x_scale,
                    'y_scale': y_scale,
                    'object_position': 2,
                }
            )
        else:
            ws.write(excel_row, 5, '', data_fmt)

    # ---- 合计行（在最后一条数据行的下一行）----
    if records:
        total_row = len(records) + 1  # 0-based（header=0, 数据=1..N, 合计=N+1）
        ws.write(total_row, 3, '合计：', total_label_fmt)
        ws.write_number(total_row, 4, total, total_money_fmt)
        ws.set_row(total_row, 40)


def create_excel_with_images(data):
    """创建带图片的Excel - xlsxwriter 方案（稳定、跨平台、不依赖 Office）
    核心：insert_image() + object_position=2 → 浮动图片（兼容所有Excel版本）
    注意：embed_image() 需要 Excel 365 2023+，老版本会显示 #VALUE!
    """
    if not HAS_XLSXWRITER:
        raise Exception("缺少 xlsxwriter，请运行: pip install xlsxwriter")
    if not HAS_PIL:
        raise Exception("缺少 Pillow，请运行: pip install Pillow")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'报销汇总_{timestamp}.xlsx'
    out_path = OUTPUT_DIR / filename

    # 过滤：只导出勾选记录（_checked 默认为 true 保持兼容）
    expense_records = data.get('expense', [])
    reimburse_records = data.get('reimburse', [])

    # 图片临时目录
    img_tmp_dir = OUTPUT_DIR / '_img_tmp'
    img_tmp_dir.mkdir(exist_ok=True)

    # 准备图片
    expense_imgs = prepare_images_xlsx(expense_records, 'expense', img_tmp_dir)
    reimburse_imgs = prepare_images_xlsx(reimburse_records, 'reimburse', img_tmp_dir)

    # 创建 Workbook
    wb = xlsxwriter.Workbook(str(out_path))

    # 费用模板
    if expense_records:
        ws1 = wb.add_worksheet('费用模板')
        write_sheet_xlsx(wb, ws1, expense_records, expense_imgs)

    # 报销模板
    if reimburse_records:
        ws2 = wb.add_worksheet('报销模板')
        write_sheet_xlsx(wb, ws2, reimburse_records, reimburse_imgs)

    # 无数据时保留一个空sheet
    if not expense_records and not reimburse_records:
        ws = wb.add_worksheet('费用模板')
        ws.write(0, 0, '无数据')

    wb.close()

    # 清理临时图片
    try:
        import shutil
        shutil.rmtree(img_tmp_dir, ignore_errors=True)
    except:
        pass

    logging.info(f"xlsxwriter 生成 Excel 成功: {out_path}")
    return out_path, filename


def create_excel_with_com(data):
    """创建带图片的Excel - 使用 COM 接口（Excel 或 WPS）实现真正嵌入单元格"""
    global OFFICE_TYPE
    
    if not HAS_WIN32COM:
        raise Exception("未检测到 Office COM 接口")
    
    if not HAS_PIL:
        raise Exception("缺少 Pillow 库")

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'报销汇总_{timestamp}.xlsx'
    out_path = OUTPUT_DIR / filename
    
    # 图片临时目录
    img_tmp_dir = OUTPUT_DIR / '_img_tmp'
    img_tmp_dir.mkdir(exist_ok=True)
    
    # 保存图片到临时目录
    def prepare_images(records, prefix):
        img_map = {}  # {row_idx: img_path}
        for idx, r in enumerate(records):
            if r.get('image'):
                try:
                    pil_img = parse_base64_image(r['image'])
                    if pil_img:
                        ext = 'png'
                        img_path = img_tmp_dir / f'{prefix}_{idx}.{ext}'
                        pil_img.save(str(img_path), format='PNG')
                        img_map[idx + 2] = str(img_path)  # Excel行号从1开始，数据从第2行开始
                except Exception as e:
                    logging.error(f"保存图片失败: {e}")
        return img_map
    
    # 准备两个表的数据
    expense_records = data.get('expense', [])
    reimburse_records = data.get('reimburse', [])
    
    expense_imgs = prepare_images(expense_records, 'expense')
    reimburse_imgs = prepare_images(reimburse_records, 'reimburse')
    
    # 启动 Excel/WPS
    import win32com.client
    app = None
    
    try:
        if OFFICE_TYPE == 'excel':
            app = win32com.client.Dispatch('Excel.Application')
        else:
            # WPS ET
            app = win32com.client.Dispatch('ET.Application')
        
        app.Visible = False
        app.DisplayAlerts = False
        
        wb = app.Workbooks.Add()
        
        # 删除默认sheet，保留需要的
        while wb.Sheets.Count > 1:
            wb.Sheets(wb.Sheets.Count).Delete()
        
        headers = ['时间', '产品', '关联项目', '发生原因', '金额', '详情截图', '是否有票', '开票主体']
        
        def write_sheet_com(ws, sheet_name, records, img_map, related_label):
            ws.Name = sheet_name
            
            # 写入表头（无背景色，黑色，加粗，居中）
            for col_idx, h in enumerate(headers, 1):
                cell = ws.Cells(1, col_idx)
                cell.Value = h
                cell.Font.Bold = True
                cell.Font.Size = 12
                cell.Font.Name = '微软雅黑'
                cell.Interior.ColorIndex = -4142  # 无背景色
                cell.Font.Color = 0x000000  # 黑色
                cell.HorizontalAlignment = -4108  # 水平居中
                cell.VerticalAlignment = -4160  # 垂直居中

            # 设置列宽（第6列=22，其他标准）
            col_widths = [12, 16, 14, 18, 12, 22, 8, 14]
            for col_idx, width in enumerate(col_widths, 1):
                try:
                    ws.Columns(col_idx).ColumnWidth = width
                    logging.info(f"设置第{col_idx}列列宽={width}")
                except Exception as e:
                    logging.warning(f"设置第{col_idx}列列宽失败: {e}，继续...")
            
            # 设置表头行高=40
            try:
                ws.Rows(1).RowHeight = 40
                logging.info("设置表头行高=40")
            except Exception as e:
                logging.warning(f"设置表头行高失败: {e}")
            
            # 第三步：写入数据并插入图片
            total = 0.0
            for row_idx, r in enumerate(records, 2):
                total += r.get('amount', 0)
                
                # 数据行样式（居中，字体10）
                def style_cell(cell):
                    cell.HorizontalAlignment = -4108  # 水平居中
                    cell.VerticalAlignment = -4160   # 垂直居中
                    cell.Font.Size = 10
                    cell.Font.Name = '微软雅黑'
                
                # 先赋值，再样式（避免被覆盖）
                cell1 = ws.Cells(row_idx, 1)
                cell1.Value = r.get('time', '')
                style_cell(cell1)
                
                cell2 = ws.Cells(row_idx, 2)
                cell2.Value = r.get('product', '')
                style_cell(cell2)
                
                cell3 = ws.Cells(row_idx, 3)
                cell3.Value = r.get(related_label, '') or '-'
                style_cell(cell3)
                
                cell4 = ws.Cells(row_idx, 4)
                cell4.Value = r.get('reason', '')
                style_cell(cell4)
                
                cell5 = ws.Cells(row_idx, 5)
                cell5.Value = r.get('amount', 0)
                cell5.NumberFormat = '¥#,##0.00'
                style_cell(cell5)
                
                cell7 = ws.Cells(row_idx, 7)
                cell7.Value = r.get('hasTicket', '')
                style_cell(cell7)
                
                cell8 = ws.Cells(row_idx, 8)
                cell8.Value = r.get('ticketEntity', '') or '-'
                style_cell(cell8)
                
                # 设置数据行行高=300
                try:
                    ws.Rows(row_idx).RowHeight = 300
                except Exception as e:
                    logging.warning(f"设置第{row_idx}行行高失败: {e}")
                
                # 插入图片（缩放匹配单元格尺寸）
                if row_idx in img_map:
                    try:
                        img_path = img_map[row_idx]
                        orig_w, orig_h = 0, 0
                        if r.get('image'):
                            try:
                                pil_img = parse_base64_image(r['image'])
                                if pil_img:
                                    orig_w, orig_h = pil_img.size
                            except:
                                pass
                        
                        # 直接用单元格尺寸
                        cell = ws.Cells(row_idx, 6)
                        cell_width = cell.Width   # 单元格宽度（磅）
                        cell_height = cell.Height  # 单元格高度（磅）
                        
                        # 计算缩放（保持比例，适应单元格）
                        if orig_w > 0 and orig_h > 0:
                            scale = min(cell_width / orig_w, cell_height / orig_h)
                            if scale > 1:
                                scale = 1  # 不放大
                            display_w = orig_w * scale
                            display_h = orig_h * scale
                        else:
                            display_w = cell_width - 4
                            display_h = cell_height - 4
                        
                        # 居中偏移
                        offset_x = (cell_width - display_w) / 2
                        offset_y = (cell_height - display_h) / 2
                        
                        pic = ws.Shapes.AddPicture(
                            img_path,
                            LinkToFile=False,
                            SaveWithDocument=True,
                            Left=cell.Left + offset_x,
                            Top=cell.Top + offset_y,
                            Width=display_w,
                            Height=display_h
                        )
                        pic.Placement = 1
                    except Exception as e:
                        logging.error(f"插入图片失败: {e}")
                        ws.Cells(row_idx, 6).Value = '图片加载失败'
            
            # 合计行
            total_row = len(records) + 2
            
            # 合计行样式
            def style_total_cell(cell):
                cell.Font.Bold = True
                cell.Font.Size = 10
                cell.Font.Name = '微软雅黑'
                cell.HorizontalAlignment = -4108  # 水平居中
                cell.VerticalAlignment = -4160   # 垂直居中
            
            style_total_cell(ws.Cells(total_row, 4))
            ws.Cells(total_row, 4).Value = '合计：'
            
            style_total_cell(ws.Cells(total_row, 5))
            ws.Cells(total_row, 5).Value = total
            ws.Cells(total_row, 5).NumberFormat = '¥#,##0.00'
            # 设置合计行行高=300
            try:
                ws.Rows(total_row).RowHeight = 300
            except Exception as e:
                logging.warning(f"设置合计行行高失败: {e}")
        
        # 写入费用模板
        if expense_records:
            if wb.Sheets.Count == 0:
                wb.Worksheets.Add()
            ws1 = wb.Worksheets(1)
            write_sheet_com(ws1, '费用模板', expense_records, expense_imgs, 'related')
        
        # 写入报销模板
        if reimburse_records:
            if wb.Sheets.Count == 1 and expense_records:
                wb.Worksheets.Add()
            ws2 = wb.Worksheets(2) if wb.Sheets.Count >= 2 else wb.Worksheets.Add()
            write_sheet_com(ws2, '报销模板', reimburse_records, reimburse_imgs, 'related')
        
        # 如果没有数据，至少一个sheet
        if not expense_records and not reimburse_records:
            ws = wb.Worksheets(1)
            ws.Name = '费用模板'
            ws.Cells(1, 1).Value = '无数据'
        
        # 保存文件
        # xlOpenXMLWorkbook = 51
        wb.SaveAs(str(out_path), 51)
        wb.Close()
        
    finally:
        if app:
            app.Quit()
    
    # 清理临时图片
    try:
        import shutil
        shutil.rmtree(img_tmp_dir, ignore_errors=True)
    except:
        pass
    
    return out_path, filename


def get_excel_creator_info():
    """获取当前Excel生成器的信息"""
    if HAS_XLSXWRITER:
        return {'type': 'xlsxwriter', 'tool': 'xlsxwriter'}
    elif HAS_WIN32COM:
        return {'type': 'com', 'tool': OFFICE_TYPE}
    else:
        return {'type': 'none', 'tool': 'none'}


# ==================== HTTP 服务器 ====================
class APIHandler(SimpleHTTPRequestHandler):
    """API处理器"""

    def _add_cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def send_json(self, data, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self._add_cors_headers()
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))

    def _read_json_body(self):
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)
        return json.loads(body.decode('utf-8'))

    def do_OPTIONS(self):
        self.send_response(200)
        self._add_cors_headers()
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        # 健康检查
        if path == '/api/health':
            self.send_json({
                'status': 'ok',
                'time': datetime.now().isoformat(),
                'excel_creator': get_excel_creator_info(),
            })

        # 加载数据
        elif path == '/api/load':
            self.send_json(load_data())

        # 最新导出文件下载
        elif path == '/api/download-latest':
            # 获取最新的导出文件
            export_files = list(OUTPUT_DIR.glob('*.xlsx'))
            if export_files:
                latest = max(export_files, key=lambda p: p.stat().st_mtime)
                self.send_file(str(latest), as_download=True)
            else:
                self.send_json({'success': False, 'error': '没有找到导出的文件'}, status=404)

        # 导出文件下载 (兼容直接访问)
        elif path.startswith('/exports/'):
            # URL解码获取文件名
            from urllib.parse import unquote
            filename = unquote(path.split('/')[-1])
            file_path = OUTPUT_DIR / filename
            if file_path.exists():
                self.send_file(str(file_path), as_download=True)
            else:
                self.send_error(404, 'File not found')

        # 静态文件
        else:
            if path == '/' or path == '':
                path = '/index.html'
            file_path = BASE_DIR / path.lstrip('/')
            
            if file_path.is_file():
                if path.endswith('.html'):
                    ct = 'text/html; charset=utf-8'
                    self.send_file(str(file_path), content_type=ct, as_download=False)
                elif path.endswith('.css'):
                    ct = 'text/css; charset=utf-8'
                    self.send_file(str(file_path), content_type=ct)
                elif path.endswith('.js'):
                    ct = 'application/javascript; charset=utf-8'
                    self.send_file(str(file_path), content_type=ct)
                else:
                    self.send_file(str(file_path))
            else:
                # SPA fallback
                self.send_file(str(BASE_DIR / 'index.html'), content_type='text/html; charset=utf-8', as_download=False)

    def send_file(self, filepath, content_type=None, as_download=False):
        """返回文件"""
        filepath = Path(filepath)
        if not filepath.exists():
            self.send_error(404, 'File not found')
            return
        
        self.send_response(200)
        ct = content_type or 'application/octet-stream'
        self.send_header('Content-Type', ct)
        
        if as_download:
            safe_name = quote(filepath.name)
            self.send_header('Content-Disposition', f'attachment; filename="{safe_name}"; filename*=UTF-8\'\'{safe_name}')
        
        self.send_header('Content-Length', filepath.stat().st_size)
        self.send_header('Cache-Control', 'no-cache')
        self._add_cors_headers()
        self.end_headers()
        
        with open(filepath, 'rb') as f:
            self.wfile.write(f.read())

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path

        try:
            if path == '/api/save':
                data = self._read_json_body()
                ok = save_data(data)
                self.send_json({'success': ok})

            elif path == '/api/export':
                data = self._read_json_body()
                # 优先使用 xlsxwriter（稳定、跨平台、不依赖 Office、图片嵌入单元格）
                # xlsxwriter insert_image + object_position=2 = Excel "放置在单元格中"
                # 回退到 COM（Excel/WPS）—— 仅在 xlsxwriter 不可用时
                if HAS_XLSXWRITER:
                    try:
                        out_path, filename = create_excel_with_images(data)
                        logging.info("使用 xlsxwriter 生成 Excel（embed_image = Place in Cell）")
                    except Exception as e:
                        logging.warning(f"xlsxwriter 生成失败，回退到 COM: {e}")
                        if HAS_WIN32COM:
                            out_path, filename = create_excel_with_com(data)
                        else:
                            raise
                elif HAS_WIN32COM:
                    out_path, filename = create_excel_with_com(data)
                    logging.info("使用 COM 接口生成 Excel")
                else:
                    self.send_json({'success': False, 'error': '无可用 Excel 生成引擎（请安装 xlsxwriter 或 Office/WPS）'}, status=500)
                    return
                self.send_json({
                    'success': True,
                    'filename': filename,
                    'download_url': f'/exports/{filename}',
                    'creator': get_excel_creator_info(),
                })

            elif path == '/api/load':
                self.send_json(load_data())

            else:
                self.send_error(404)
        except Exception as e:
            logging.error(f"处理请求出错: {path} → {e}", exc_info=True)
            self.send_json({'success': False, 'error': str(e)}, status=500)

    def log_message(self, format, *args):
        if args and 'favicon' in str(args[0]):
            return
        logging.info(args[0] if args else format)


# ==================== 主入口 ====================
def main():
    load_config()
    init_logging()

    # 检测 Office 工具 (Excel / WPS)
    detect_office_tool()
    logging.info(f"Excel生成器: {get_excel_creator_info()}")

    if not HAS_OPENPYXL or not HAS_PIL:
        logging.error("缺少必要依赖！")
        print("\n❌ 缺少必要依赖！请先安装：")
        print("   pip install openpyxl Pillow\n")
        input("按回车键退出...")
        sys.exit(1)

    # 单实例检测
    old_pid = check_single_instance()
    if old_pid:
        print(f"\n⚠️  程序已在运行中（PID: {old_pid}）")
        print(f"   请关闭后重试，或前往 http://localhost:{config['port']} 使用")
        input("按回车键退出...")
        sys.exit(0)

    # 端口检测
    start_port = config['port']
    port = find_available_port(start_port)
    if port is None:
        print("❌ 无法找到可用端口")
        sys.exit(1)
    
    if port != start_port:
        print(f"⚠️  端口 {start_port} 已被占用，自动使用端口 {port}")

    host = "0.0.0.0"
    server = HTTPServer((host, port), APIHandler)

    local_ip = get_local_ip()
    local_url = f"http://localhost:{port}"
    lan_url = f"http://{local_ip}:{port}"

    # 打印启动信息
    H = '=' * 48
    print()
    print(f'  {H}')
    print(f'  |{" " * 46}|')
    print(f'  |    报销费用填写工具 v1.0                    |')
    print(f'  |    Expense & Reimbursement Tool             |')
    print(f'  |{" " * 46}|')
    print(f'  {H}')
    print()
    print(f'    Author : aigc创意人竹相左边')
    print(f'    Engine : Python + xlsxwriter')
    print()
    print(f'    {"-" * 44}')
    print(f'    Status : Running')
    print(f'    Local  : {local_url}')
    print(f'    LAN    : {lan_url}')
    print(f'    Data   : {DATA_DIR}')
    print(f'    {"-" * 44}')
    print()
    print(f'    Stop: Ctrl+C')
    print()

    # 自动打开浏览器
    def open_browser():
        time.sleep(1.5)
        webbrowser.open(local_url)

    threading.Thread(target=open_browser, daemon=True).start()

    # 优雅退出
    def shutdown(signum, frame):
        print("\n正在停止服务...")
        # 清理锁文件
        lock_file = DATA_DIR / 'server.lock'
        if lock_file.exists():
            lock_file.unlink()
        logging.info("服务关闭")
        server.shutdown()
        sys.exit(0)

    signal.signal(signal.SIGTERM, shutdown)
    signal.signal(signal.SIGINT, shutdown)

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        shutdown(None, None)


# ---------------------------------------------------------------------------
# 图片真正的"放置在单元格中"：xlsx XML 后处理
# ---------------------------------------------------------------------------
import zipfile, re, shutil, os, tempfile

def _embed_images_into_cells(xlsx_path):
    """
    xlsxwriter 生成的图片默认是 twoCellAnchor（浮动，不随单元格缩放）。
    此函数将其改为 oneCellAnchor，使图片真正嵌入单元格。
    """
    tmp_dir = tempfile.mkdtemp(prefix='xlsx_embed_')
    try:
        # 解压 xlsx
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            z.extractall(tmp_dir)

        # 遍历所有 sheet 的 XML
        xl_dir = os.path.join(tmp_dir, 'xl')
        drawings_dir = os.path.join(xl_dir, 'drawings')
        if not os.path.isdir(drawings_dir):
            return

        for drawing_file in os.listdir(drawings_dir):
            if not drawing_file.endswith('.xml'):
                continue
            drawing_path = os.path.join(drawings_dir, drawing_file)

            with open(drawing_path, 'r', encoding='utf-8') as f:
                xml_content = f.read()

            original = xml_content

            # xlsxwriter 生成格式：<xdr:twoCellAnchor editAs="oneCell">...<xdr:from>...</xdr:from><xdr:to>...</xdr:to>...</xdr:twoCellAnchor>
            # 目标格式：<xdr:oneCellAnchor>...<xdr:from>...</xdr:from>...</xdr:oneCellAnchor>（无 to 节点）
            
            # 1. 先把带属性的开标签统一处理（如 editAs="oneCell"）
            xml_content = re.sub(r'<xdr:twoCellAnchor[^>]*>', '<xdr:oneCellAnchor>', xml_content)
            # 2. 闭标签
            xml_content = xml_content.replace('</xdr:twoCellAnchor>', '</xdr:oneCellAnchor>')
            # 3. 去掉 <xdr:to>...</xdr:to> 节点（oneCellAnchor 不需要 to）
            xml_content = re.sub(r'<xdr:to>.*?</xdr:to>', '', xml_content, flags=re.DOTALL)

            if xml_content != original:
                with open(drawing_path, 'w', encoding='utf-8') as f:
                    f.write(xml_content)
                logging.info(f"  图片已改为单元格嵌入: {drawing_file}")

        # 重新打包 xlsx
        tmp_xlsx = xlsx_path + '.tmp'
        with zipfile.ZipFile(tmp_xlsx, 'w', compression=zipfile.ZIP_DEFLATED) as z:
            for root, dirs, files in os.walk(tmp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, tmp_dir)
                    z.write(file_path, arcname)
        os.replace(tmp_xlsx, xlsx_path)

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


if __name__ == '__main__':
    main()
